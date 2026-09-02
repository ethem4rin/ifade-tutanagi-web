# -*- coding: utf-8 -*-
"""İşveren için işçi bilgi formu: boş Excel üret + dolu Excel oku.

İşveren teftişten önce doldurur; program içe aktarınca işçi kimlikleri hazır gelir.
İfade gövdesi (yeşil/kırmızı) yine müfettiş tarafından girilir."""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (excel_baslik, isci_sozluk_anahtari)
SUTUNLAR = [
    ("Sıra", None),
    ("Adı Soyadı", "ad_soyad"),
    ("T.C. Kimlik No", "tc"),
    ("Doğum Yeri", "dogum_yeri"),
    ("Doğum Yılı", "dogum_yili"),
    ("Baba Adı", "baba"),
    ("Ana Adı", "ana"),
    ("Görevi", "gorevi"),
    ("Telefon No", "telefon"),
    ("İkametgah (Açık Adres)", "ikametgah"),
]

_KENAR = Border(*[Side(style="thin", color="BBBBBB")] * 4)


def form_olustur(isyeri: dict, cikti_yol: str, satir_sayisi: int = 25) -> str:
    """Seçili işyeri için boş işçi bilgi formu (xlsx) üretir."""
    wb = Workbook()
    ws = wb.active
    ws.title = "İşçi Bilgileri"

    # üst bilgi
    ws["A1"] = "İŞÇİ BİLGİ FORMU (İfade Tutanağı için)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    ws["A2"] = f"İşyeri: {isyeri.get('isveren', '')}"
    ws["A3"] = f"Sicil No: {isyeri.get('sicil', '')}"
    for r in (2, 3):
        ws[f"A{r}"].font = Font(bold=True, size=11)
        ws.merge_cells(f"A{r}:J{r}")
    ws["A4"] = ("Lütfen her işçi için aşağıdaki bilgileri eksiksiz doldurunuz. "
                "Doğum yeri/yılı ve ana-baba adı kimlik kartındaki gibi yazılmalıdır.")
    ws["A4"].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells("A4:J4")

    # başlık satırı
    bas = 6
    baslik_dolgu = PatternFill("solid", fgColor="1F6AA5")
    for j, (et, _k) in enumerate(SUTUNLAR, 1):
        c = ws.cell(row=bas, column=j, value=et)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = baslik_dolgu
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _KENAR

    # boş satırlar + sıra no
    for i in range(satir_sayisi):
        r = bas + 1 + i
        ws.cell(row=r, column=1, value=i + 1).alignment = Alignment(horizontal="center")
        for j in range(1, len(SUTUNLAR) + 1):
            ws.cell(row=r, column=j).border = _KENAR

    # sütun genişlikleri
    genislik = [6, 24, 18, 16, 10, 16, 16, 22, 16, 40]
    for j, g in enumerate(genislik, 1):
        ws.column_dimensions[get_column_letter(j)].width = g
    ws.freeze_panes = f"A{bas + 1}"

    os.makedirs(os.path.dirname(os.path.abspath(cikti_yol)), exist_ok=True)
    wb.save(cikti_yol)
    return cikti_yol


def form_oku(xlsx_yol: str):
    """Dolu formu okuyup işçi sözlüğü listesi döndürür (tutanak alanlarına eşlenmiş)."""
    wb = load_workbook(xlsx_yol, data_only=True)
    ws = wb.active
    # başlık satırını bul
    bas = None
    for r in range(1, 15):
        if str(ws.cell(row=r, column=2).value).strip() == "Adı Soyadı":
            bas = r
            break
    if bas is None:
        bas = 6
    anahtarlar = [k for _et, k in SUTUNLAR]

    isciler = []
    for r in range(bas + 1, ws.max_row + 1):
        ham = {}
        for j, k in enumerate(anahtarlar, 1):
            if k is None:
                continue
            v = ws.cell(row=r, column=j).value
            ham[k] = ("" if v is None else str(v)).strip()
        if not ham.get("ad_soyad"):
            continue
        isciler.append(_esle(ham))
    return isciler


def _esle(h: dict) -> dict:
    """Ham form sütunlarını tutanak alan adlarına çevir."""
    dy = " / ".join(x for x in (h.get("dogum_yeri", ""), h.get("dogum_yili", "")) if x)
    ba = "/ ".join(x for x in (h.get("baba", ""), h.get("ana", "")) if x)
    return {
        "ad_soyad": h.get("ad_soyad", ""),
        "tc": h.get("tc", ""),
        "dogum_yeri_yili": dy,
        "baba_ana": ba,
        "gorevi": h.get("gorevi", ""),
        "telefon": h.get("telefon", ""),
        "ikametgah": h.get("ikametgah", ""),
    }
